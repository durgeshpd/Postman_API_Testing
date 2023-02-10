import re
import openpyxl
from openpyxl.utils import get_column_letter

with open("log.txt", "r") as log:
    text = log.read()

with open("request_log.txt", "r") as request_log:
    request_text = request_log.read()

wb = openpyxl.Workbook()
ws = wb.active

ws.cell(row=1, column=1, value="S.No.")
ws.cell(row=1, column=2, value="Test count")
ws.cell(row=1, column=3, value="Test time")
ws.cell(row=1, column=4, value="Source/Loader")
ws.cell(row=1, column=5, value="Firmware Version")
ws.cell(row=1, column=6, value="WDT CTL")
ws.cell(row=1, column=7, value="WDT ALT")
ws.cell(row=1, column=8, value="WDT RST")
ws.cell(row=1, column=9, value="API Message")
ws.cell(row=1, column=10, value="Status of Update")
ws.cell(row=1, column=11, value="Flag")

for i in range(1, 12):
    if i in [1, 2, 6, 7, 8, 11]:
        ws.column_dimensions[get_column_letter(i)].width = 10
    elif i in [4, 5]:
        ws.column_dimensions[get_column_letter(i)].width = 15
    else:
        ws.column_dimensions[get_column_letter(i)].width = 20


upgraded_list = re.findall(r"ROLL BACK Completed", text)
src_flag_list = re.findall(r"Source_Flag:(.*)", text)
ldr_flag_list = re.findall(r"Loader_Flag : (.*)", text)
ver_list = re.findall(r"ver 1\.0|ver 1\.1", text)
api_message_list = re.findall(r"api message...: (start_update code: 81)", text)
wdt_control_list = re.findall(r"WDT register state (.*)", text)
wdt_rst_list = re.findall(r"WDT RSTCNT:(.*)", text)
wdt_alt_list = re.findall(r"WDT ALT:(.*)", text)

test_count_list = re.findall(r"^(\d+)", request_text, re.MULTILINE)
test_time_list = re.findall(r"\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}", request_text)

upgraded_len = len(upgraded_list)
src_flag_len = len(src_flag_list)
ldr_flag_len = len(ldr_flag_list)
ver_len = len(ver_list)
api_message_len = len(api_message_list)
wdt_control_len = len(wdt_control_list)
wdt_alt_len = len(wdt_alt_list)
wdt_rst_len = len(wdt_rst_list)
test_time_len = len(test_time_list)
test_count_len = len(test_time_list)

lists = [ver_list, wdt_control_list, wdt_alt_list, ldr_flag_list, wdt_rst_list, api_message_list, test_count_list, test_time_list]
max_len = max(len(l) for l in lists)

for i in range(max_len):
    row = (i + 1) * 4
    ws.cell(row=row, column=1, value=i + 1)
    ws.cell(row=row, column=5, value=ver_list[i] if i < len(ver_list) else '')
    ws.cell(row=row, column=6, value=wdt_control_list[i] if i < len(wdt_control_list) else '')
    ws.cell(row=row, column=7, value=wdt_alt_list[i] if i < len(wdt_alt_list) else '')
    ws.cell(row=row, column=8, value=wdt_rst_list[i] if i < len(wdt_rst_list) else '')
    ws.cell(row=row, column=9, value=api_message_list[i] if i < len(api_message_list) else '')
    ws.cell(row=row, column=10, value=upgraded_list[i] if i < len(upgraded_list) else '')
    ws.cell(row=row, column=11, value=src_flag_list[i] if i < len(src_flag_list) else '')
    ws.cell(row=row, column=4, value='Source')
    ws.cell(row=row + 1, column=4, value='Loader')
    ws.cell(row=row + 2, column=4, value='Loader')
    ws.cell(row=row + 3, column=4, value='Loader')
    for j in range(3):
        ws.cell(row=row + j + 1, column=11, value=ldr_flag_list[j + i * 3] if j + i * 3 < len(ldr_flag_list) else '')
    ws.cell(row=row, column=2, value=test_count_list[i] if i < len(test_count_list) else '')
    ws.cell(row=row, column=3, value=test_time_list[i] if i < len(test_time_list) else '')

wb.save("log_report.xlsx")
